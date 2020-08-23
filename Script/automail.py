import time, os, sys,smtplib, ssl, threading
from openpyxl import load_workbook
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import tkinter as tk
from tkinter import filedialog, Entry, scrolledtext

def resource_path(relative_path):
    """
    Deals with external files that will be neaded after
    packaging.
    :param relative_path: path of file.
    :return: relative path.
    """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def authenticate():
    """
    Tests user authentification on the outlook server.
    :return: Updates Global variable Login that will
    be checked before starting to send actual mails.
    """
    cmd("Attempting Login:")
    # Preventing further attempt to start.
    startButton['state'] = tk.DISABLED
    # Email Configuration.
    SMTP_SERVER = "smtp-mail.outlook.com"
    PORT = 587  # for TLS on OUTLOOK.
    # Create a secure SSL context
    context = ssl.create_default_context()
    # Try to log in to server and send email
    global Login
    try:
        server = smtplib.SMTP(SMTP_SERVER,PORT)
        # Identifying user on server.
        server.ehlo()
        # Secure the connection
        server.starttls(context=context)
        # Identifying user on server.
        server.ehlo()
        server.login(UserEmail.get(), UserpasswD.get())
        cmd("Login Successful")
        Login = True
        StartProg()
    except Exception as e:
        # Print any error messages to stdout
        cmd("Login Failed!")
        cmd("Check email and password and try again.")
        Login = False
        startButton['state'] = tk.NORMAL
    finally:
        server.quit()

def StartProg():
    """
    Main Program that will deal with basically all the backend stuff.
    :return: None.
    """
    # Checking if Loging is True or False to see in password entered
    # is correct.
    if Login != False:

        cmd("Clearing Console")
        time.sleep(1)
        clearCMD()
        DATA = update_dict(EXCEL_PATH)
        num = len(DATA['NAME'])
        for i in range(num):
            cmd(f"SENDING to {DATA['NAME'][i]}...")
            message_ready = msg_build(DATA['NAME'][i]\
                ,DATA['SURNAME'][i],DATA['MARK'][i],DATA["EMAIL"][i],\
                    UserHeader.get('1.0', 'end-1c'), TotMark.get(), UserSubJ.get(),\
                        UserEmail.get(), UserName.get(), UserTitle.get())
            send_mail(DATA['EMAIL'][i], message_ready,\
                UserEmail.get(), UserpasswD.get())
            cmd(f"SENT TO {DATA['NAME'][i]}!")
            time.sleep(1) # Outlook can send only 30 email/min in 365 account.
        cmd("ALL Done :)")
        startButton['state'] = tk.NORMAL

def update_dict(EXCEL_PATH):
    """
    Generated Dictionary to work with.

    Function that reads from the excel file and updates a dictionary
    that we will then further use in our program.

    Args: None
    Returns: None

    Updates Dictionary.
    """
    # Data Dictionary for further processing.
    DATA = {'NAME': [], 'SURNAME': [], 'EMAIL': [], 'MARK': []}
    # Workbook object is created to open.
    WB_OBJ = load_workbook(filename=EXCEL_PATH)
    # Sheet Object.
    SHEET_OBJ = WB_OBJ.active
    row_len = SHEET_OBJ.max_row
    col_len = SHEET_OBJ.max_column

    # Finding the correct columns that we'll be working with.
    for colIndex in range(1,col_len + 1):
        cell_obj = SHEET_OBJ.cell(row = 1, column = colIndex)
        # Finding the Name Column.
        if cell_obj.value == "Name":
            name_col = colIndex
        # Finding the Surname Column.
        if cell_obj.value == "Surname":
            surname_col = colIndex
        # Finding the Email Column.
        if cell_obj.value == "Email":
            email_col = colIndex
        # Finding the Mark Column.
        if cell_obj.value == "Mark":
            mark_col = colIndex

    # Reading the Name column of the excel sheet. (column name_col)
    for rowIndex in range(2,row_len + 1):
        cell_obj = SHEET_OBJ.cell(row = rowIndex, column = name_col)
        DATA['NAME'].append(cell_obj.value)
    # Reading the surname column of the excel sheet.
    for rowIndex in range(2,row_len + 1):
        cell_obj = SHEET_OBJ.cell(row = rowIndex, column = surname_col)
        DATA['SURNAME'].append(cell_obj.value)
    # Reading the email column of the excel sheet.
    for rowIndex in range(2,row_len + 1):
        cell_obj = SHEET_OBJ.cell(row = rowIndex, column = email_col)
        DATA['EMAIL'].append(cell_obj.value)
    # Reading the mark column of the excel sheet.
    for rowIndex in range(2,row_len + 1):
        cell_obj = SHEET_OBJ.cell(row = rowIndex, column = mark_col)
        DATA['MARK'].append(cell_obj.value)
    return DATA

def msg_build(name,surname,mark,receiver_mail,header_message,total_mark, SUBJECT,\
    SENDER_EMAIL, SENDER_NAME,title):
    """
    Functions that builds the custom message for each student.

    :param name: Name of student.
    :param surname: Surname of student.
    :param mark: Mark of student.
    :param receiver_mail: Email Address of the student.
    :param header_message: Static header message.
    :param total_mark: The total mark of the test.
    :return: Custom Message to be send to this student.
    """
    message = MIMEMultipart("alternative")
    message["Subject"] = SUBJECT
    message["From"] = SENDER_EMAIL
    message["TO"] = receiver_mail
    # insert your html content below(The static message.
    # Everything in curly brackets are variables that will be
    # completed by the program.)
    # It comes with an example of Mr Kewalnath Seereekissoon.
    tab_syn ="""
        border: 1px solid black;
        border-collapse: collapse;
        text-align: center;
"""
    html =(f"""
<html>
    <head>
        <style>
        table, th, td {{{tab_syn}}}
        </style>
    </head>
    <body>
        <p>
            Hello <strong>{name}</strong>,<br><br>
            {header_message}<br><br>
            <table style="width:75%">
                <tr>
                    <th>Name</th><th>Total({total_mark} Marks)</th> 
                </tr>
                <tr>
                    <td><strong>{surname}</strong> {name}</td><td>{mark}</td>
                </tr>
            </table>
          </p><br>
        <p><strong>Kind Regards,</strong><br></p>
        <p style="color:orange"></style>{SENDER_NAME}</p>
        <p>{title}</p>
    </body>
</html>
""")

    #txt = MIMEText(text,"plain")
    htm = MIMEText(html, "html")
    #message.attach(txt)
    message.attach(htm)
    msg_ready = message.as_string()
    return msg_ready

def send_mail(receiver_mail, message, SENDER_EMAIL, PASSWORD):
    # Email Configuration.
    SMTP_SERVER = "smtp-mail.outlook.com"
    PORT = 587  # for TLS on OUTLOOK.
    # Create a secure SSL context
    context = ssl.create_default_context()
    # Try to log in to server and send email
    try:
        server = smtplib.SMTP(SMTP_SERVER,PORT)
        # Identifying user on server.
        server.ehlo()
        # Secure the connection
        server.starttls(context=context)
        # Identifying user on server.
        server.ehlo()
        server.login(SENDER_EMAIL, PASSWORD)
        server.sendmail(SENDER_EMAIL, receiver_mail, message)
    except Exception as e:
        # Print any error messages to stdout
        cmd(e)
    except:
        cmd("An error Occured!")
    finally:
        server.quit()

def loadExcelFunc():
    """
    Open System file browser dialog to select the .xlsx file.
    :return: None

    Update the path of the excel file globally( EXCEL_PATH).
    """
    filename = filedialog.askopenfilename(initialdir="/", title="Select .xlsx File",\
        filetypes=(("Excel", "*.xlsx"), ("All Files", "*.*")))
    global EXCEL_PATH
    EXCEL_PATH = filename
    if EXCEL_PATH != "":
        cmd(f"Loaded: {EXCEL_PATH}")


def cmd(command_output):
    """
    Function to handle outputting to the user Console output frame.

    :param command_output: String That we want to be output to the user.
    :return: None

    Displays in the console output frame.
    """
    global commandLineCount, consoleOutputLabel
    ttyy = (f"> {command_output}")
    if len(ttyy) > 35:
        font_size = 10
    else:
        font_size = 12
    consoleOutputLabel = tk.Label(console_frame, text = ttyy, fg="white", bg="gray",\
                                  font=("terminal", font_size))
    consoleOutputLabel.pack()
    commandLineCount += 1
    if commandLineCount > 17:
        clearCMD()

def clearCMD():
    """
    Clears the Console Output frame.
    :return: None
    """
    global commandLineCount
    print(commandLineCount)
    for widget in console_frame.winfo_children():
        widget.destroy()
    consoleOutputLabel = tk.Label(console_frame, text = "Console Output:", \
        fg="white", bg="gray", font=("terminal", 12))
    consoleOutputLabel.pack()
    commandLineCount = 2

def helpd(field):
    """
    :param field: Which field's help button was pressed.
    :return: None
    Pass on the short help instruction to the cmd function that will then be seen by
    the user.
    """
    if field == "name":
        cmd("Name Field: Enter your First and Last name in this field.     ")
    if field == "title":
        cmd("Title Field: Enter your Job Title in this field.              ")
    if field == "emailaddress":
        cmd("Email Address Field: Enter your Email Address in this field.  ")
    if field == "emailsubj":
        cmd("Email Subject: Enter your Email Subject in this field.        ")
    if field == "totmark":
        cmd("Total Mark: Enter total assignment mark in this field.        ")
    if field == "emailheader":
        cmd("Email Header: Enter your Email Header in this field.          ")
        cmd("Note: Next line will not be applied on the mail. Everything   ")
        cmd("will be in a single line.")
    if field == "importExcel":
        cmd("Import Excel Button: Select your Excel File.                  ")
    if field == "password":
        cmd("Password Field: Enter your password here.                     ")
        cmd("All characters are hidden for security purpose.               ")
    

# Global Count for command line     
commandLineCount = 2

### MAIN GUI STARTS HERE!! ###

root = tk.Tk()
root.title("Batch Mailing Python Automated")
canvas = tk.Canvas(root, height=768, width=1024, bg="white").pack()

# Top Frame + Logo.
top_frame = tk.Frame(root, bg="yellow")
top_frame.place(relwidth=0.9, relheight=0.3, relx=0.05, rely=0)
poly_logo_path = tk.PhotoImage(file=resource_path("some_back.png"))
poly_top_logo = tk.Label(top_frame, image = poly_logo_path).pack(fill="both", expand="yes")

# Used frame to get the user inputs(used grid.)
bottom_frame = tk.Frame(root, bg="white")
bottom_frame.place(relwidth=0.5, relheight=0.6, relx=0.01, rely=0.3)

# User Name Input.
NameLabel = tk.Label(bottom_frame,text="Name:", fg="black",bg="white",\
    font=("Arial", 12)).grid(row=0)
UserName = Entry(bottom_frame, width=30, font=("Arial", 12))
UserName.grid(column=3, row=0, pady=10)
# We get the input here using ## UserName.get() ## normally in function of start.

# User Title Input.
TitleLabel = tk.Label(bottom_frame,text="Job Title:", fg="black",bg="white",\
    font=("Arial", 12)).grid(row=1)
UserTitle = Entry(bottom_frame, width=30, font=("Arial", 12))
UserTitle.grid(column=3, row=1)

# User Email Input.
EmailLabel = tk.Label(bottom_frame,text="Email ADDRESS:", fg="black",bg="white",\
    font=("Arial", 12)).grid( row=2)
UserEmail = Entry(bottom_frame, width=30, font=("Arial", 12))
UserEmail.grid(column=3, row=2, pady=10)

# User Email Subject.
SubJLabel = tk.Label(bottom_frame,text="Email SUBJECT:", fg="black",bg="white",\
    font=("Arial", 12)).grid(row=3)
UserSubJ = Entry(bottom_frame, width=30, font=("Arial", 12))
UserSubJ.grid(column=3, row=3, pady=10)

# Email Header.
HeaderLabel = tk.Label(bottom_frame,text="Email HEADER:", fg="black",bg="white",\
    font=("Arial", 12)).grid(row=4)
UserHeader = scrolledtext.ScrolledText(bottom_frame, wrap=tk.WORD, width=33, height=6)
UserHeader.grid(column=3, row=4, pady=10)

# Total Mark Input.
TotMarkLabel = tk.Label(bottom_frame,text="Total MARK:", fg="black",bg="white",\
    font=("Arial", 12)).grid( row=5)
TotMark = Entry(bottom_frame, width=30, font=("Arial", 12))
TotMark.grid(column=3, row=5, pady=10)

# Excel File path.
ExcelPathLabel = tk.Label(bottom_frame,text="Import Excel File:", fg="black",bg="white",\
    font=("Arial", 12)).grid(row=6)

## Excel Load Button.
loadExcel = tk.Button(bottom_frame, text="Load Excel", padx=40,\
    command= lambda: threading.Thread(target=loadExcelFunc).start())
loadExcel.grid(column=3, row=6, padx=40)

# User Email Password.
passwDLabel = tk.Label(bottom_frame,text="Password:", fg="black",bg="white",\
    font=("Arial", 12)).grid(row=7)
UserpasswD = Entry(bottom_frame,show="*", width=30, font=("Arial", 12))
UserpasswD.grid(column=3, row=7, pady=10)

# Start button.
startButton= tk.Button(bottom_frame, text="START", font=("Arial",15),\
    command= lambda: threading.Thread(target=authenticate).start())
startButton.grid(column=3, row=8)


# Console Output Frame.
console_frame = tk.Frame(root, bg="gray")
console_frame.place(relwidth=0.50, relheight=0.5, relx=0.48, rely=0.31)
consoleOutputLabel = tk.Label(console_frame, text = "Powered by PYTHON.", \
    fg="white", bg="gray", font=("terminal", 12))
consoleOutputLabel.pack()
consoleOutputLabel = tk.Label(console_frame, text = "Console Output:", \
    fg="white", bg="gray", font=("terminal", 12))
consoleOutputLabel.pack()

# Help Buttons:
## Name Field:
UNameHelp = tk.Button(bottom_frame, text="?", font=("Arial",10), \
    command= lambda: helpd("name") )
UNameHelp.grid(column=4, row=0)
## Title Field:
UTitleHelp = tk.Button(bottom_frame, text="?", font=("Arial",10), \
    command= lambda: helpd("title") )
UTitleHelp.grid(column=4, row=1)
## Address Field:
UAddressHelp = tk.Button(bottom_frame, text="?", font=("Arial",10), \
    command= lambda: helpd("emailaddress") )
UAddressHelp.grid(column=4, row=2)

## Subject Field:
USubJHelp = tk.Button(bottom_frame, text="?", font=("Arial",10), \
    command= lambda: helpd("emailsubj") )
USubJHelp.grid(column=4, row=3)

## Header Field.
UHeaderHelp = tk.Button(bottom_frame, text="?", font=("Arial",10), \
    command= lambda: helpd("emailheader") )
UHeaderHelp.grid(column=4, row=4, padx=10)

## Total Mark Field.
UHeaderHelp = tk.Button(bottom_frame, text="?", font=("Arial",10), \
    command= lambda: helpd("totmark") )
UHeaderHelp.grid(column=4, row=5)

## Import Excel Button::
ImpExcelHelp = tk.Button(bottom_frame, text="?", font=("Arial",10), \
    command= lambda: helpd("importExcel") )
ImpExcelHelp.grid(column=4, row=6, padx= 20)

## Password Field:
UPassHelp = tk.Button(bottom_frame, text="?", font=("Arial",10), \
    command= lambda: helpd("password") )
UPassHelp.grid(column=4, row=7)

root.mainloop()
