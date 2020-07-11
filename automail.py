import smtplib, ssl, time, os, getpass
from openpyxl import load_workbook
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Email Configuration.
SMTP_SERVER = "smtp-mail.outlook.com"
PORT = 587 # for TLS on OUTLOOK.

# Data Dictionary for further processing.
DATA = {'NAME':[],'SURNAME':[],'EMAIL':[],'MARK':[]}

def update_dict():
    """
    Generated Dictionary to work with.

    Function that reads from the excel file and updates a dictionary
    that we will then further use in our program.

    Args: None
    Returns: None

    Updates Dictionary.
    """
    row_len = SHEET_OBJ.max_row
    # Reading the first column of the excel sheet. (column 1)
    for rowIndex in range(2,row_len + 1):
        cell_obj = SHEET_OBJ.cell(row = rowIndex, column = 1)
        DATA['NAME'].append(cell_obj.value) 
    # Reading the second column of the excel sheet. (column 2)
    for rowIndex in range(2,row_len + 1):
        cell_obj = SHEET_OBJ.cell(row = rowIndex, column = 2)
        DATA['SURNAME'].append(cell_obj.value)
    # Reading the third column of the excel sheet. (column 3)
    for rowIndex in range(2,row_len + 1):
        cell_obj = SHEET_OBJ.cell(row = rowIndex, column = 3)
        DATA['EMAIL'].append(cell_obj.value)
    # Reading the forth column of the excel sheet. (column 4)
    for rowIndex in range(2,row_len + 1):
        cell_obj = SHEET_OBJ.cell(row = rowIndex, column = 4)
        DATA['MARK'].append(cell_obj.value)

def msg_build(name,surname,mark,receiver_mail,header_message,total_mark):
    """
    Functions that builds the custom message for each student.

    :param name: Name of student.
    :param surname: Surname of student.
    :param mark: Mark of student.
    :param receiver_mail: Email Address of the student.
    :param header_message: Static header message.
    :param total_mark: The total mark of the test.
    :return: Custom Message to be send to this student.
    """
    message = MIMEMultipart("alternative")
    message["Subject"] = SUBJECT
    message["From"] = SENDER_EMAIL
    message["TO"] = receiver_mail
    # insert your html content below(The static message. 
    # Everything in curly brackets are variables that will be
    # completed by the program.)
    tab_syn ="""
        border: 1px solid black;
        border-collapse: collapse;
        text-align: center;
"""
    html =(f"""
<html>
    <head>
        <style>
        table, th, td {{{tab_syn}}}
        </style>
    </head>
    <body>
        <p>
            Hello <strong>{name}</strong>,<br><br>
            {header_message}<br><br>
            <table style="width:75%">
                <tr>
                    <th>Name</th><th>Total({total_mark} Marks)</th> 
                </tr>
                <tr>
                    <td><strong>{surname}</strong> {name}</td><td>{mark}</td>
                </tr>
            </table>
          </p><br>
       <p>Thanks & Regards,<br><br>
       <strong>YOUR NAME HERE</strong><br>
       TITLE here!<br><br>
       <strong>Email:</strong> email address here!<br>
       <strong>Github:</strong>
       <a href="https://www.github.com/GirishMahabir">github.com/GirishMahabir</a><br>
       <strong>This was an automated python script.</strong>
    </p>
    </body>
</html>
""")

    #txt = MIMEText(text,"plain")
    htm = MIMEText(html, "html")
    #message.attach(txt)
    message.attach(htm)
    msg_ready = message.as_string()
    return msg_ready

def send_mail(receiver_mail, message):
    # Create a secure SSL context
    context = ssl.create_default_context()
    # Try to log in to server and send email
    try:
        server = smtplib.SMTP(SMTP_SERVER,PORT)
        # Identifying user on server.
        server.ehlo()
        # Secure the connection
        server.starttls(context=context) 
        # Identifying user on server.
        server.ehlo()
        server.login(SENDER_EMAIL, PASSWORD)
        server.sendmail(SENDER_EMAIL, receiver_mail, message)
    except Exception as e:
        # Print any error messages to stdout
        print(e)
    finally:
        server.quit()

def main():
    # Accessing GLOBAL VARIABLES.
    global SENDER_EMAIL, PASSWORD, FILE_PATH, SHEET_OBJ, SUBJECT
    clear = lambda: os.system('cls')
    print("WELCOME TO OUR BATCH AUTOMATED EMAIL SCRIPT.\n")
    print("WARNING: For the next Prompt:\nIF script and excel file are in \
the same folder you can just put the file name along with the correct extention.\n")
    # Path of Excel File.
    FILE_PATH = input("Enter your .xlsx file PATH along with the correct extention: ")
    # Workbook object is created to open.
    WB_OBJ = load_workbook(filename=FILE_PATH)
    # Sheet Object.
    SHEET_OBJ = WB_OBJ.active
    # Taking sender email.
    SENDER_EMAIL = input("Enter your email address: ")
    # Taking the email subject.
    SUBJECT = input("Enter your email Subject: ")
    # Taking Message top part.
    header_message = input("Enter Message you want to keep static above the mark table: ")
    # Taking Total Mark of the test.
    total_mark = input("The test was on how much marks? ")
    # Taking user password with hidden characters.
    PASSWORD = getpass.getpass(prompt= f"Please enter you password for {SENDER_EMAIL} HERE: ")
    clear() # Clearing terminal.
    print("Cleared Terminal to hide sensitive information.\n[STARTING...]")
    update_dict() # Calling the update_dict function and updating the dictionary.
    num = len(DATA['NAME'])
    for i in range(num):
        print(f"SENDING to {DATA['NAME'][i]}...")
        message_ready = msg_build(DATA['NAME'][i]\
            ,DATA['SURNAME'][i],DATA['MARK'][i],DATA["EMAIL"][i], header_message, total_mark)
        send_mail(DATA['EMAIL'][i], message_ready)
        print(f"SENT TO {DATA['NAME'][i]}!")
        time.sleep(1) # Waiting a bit to not overload the server.
    print("ALL Done, [EXITING]...")

if __name__=="__main__":
    main()
